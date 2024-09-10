import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Alignment, Font,  PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

def create_overview_sheet(output_file_path):
    """
    Create an 'Overview' sheet in the specified Excel file with information about
    the total number of countries, total fee per each country, and total fee per each year.

    Parameters:
    - output_file_path (str): The file path of the Excel file where the results are saved.
    """
    # Load the existing workbook
    workbook = load_workbook(output_file_path)
    
    # Load the results sheet into a DataFrame
    results_df = pd.read_excel(output_file_path, sheet_name=workbook.sheetnames[0])
    
    # Create a new sheet for the overview
    if 'Overview' in workbook.sheetnames:
        overview_sheet = workbook['Overview']
    else:
        overview_sheet = workbook.create_sheet(title='Overview')
    
    # Total fee per each country
    total_fees_per_country = results_df.groupby('Publication Country')['Total Fees'].sum().reset_index()
    
    # Total fee per each year
    year_columns = [col for col in results_df.columns if col.isdigit()]
    total_fees_per_year = results_df[year_columns].sum().reset_index()
    total_fees_per_year.columns = ['Year', 'Maintenance Cost ($)']

    # Define styles for headers
    header_font = Font(bold=True, color="000000")  # Black text
    header_fill = PatternFill("solid", fgColor="FFFFFF")  # White background
    header_alignment = Alignment(horizontal='center', vertical='center')


    for row in total_fees_per_country.itertuples(index=False):
        overview_sheet.append(row)
    
    # Add some space before the next table
    overview_sheet.append([])

    # Write the second table header
    overview_sheet.append(['Year', 'Maintenance Cost ($)'])
    for cell in overview_sheet[overview_sheet.max_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for row in total_fees_per_year.itertuples(index=False):
        overview_sheet.append(row)
    
    # Match column width and row height to the main sheet
    main_sheet = workbook[workbook.sheetnames[0]]
    for col in range(1, overview_sheet.max_column + 1):
        col_letter = get_column_letter(col)
        overview_sheet.column_dimensions[col_letter].width = 15  # Adjusting the width for better visibility

    # Set row height for the header rows
    overview_sheet.row_dimensions[1].height = main_sheet.row_dimensions[1].height
    overview_sheet.row_dimensions[overview_sheet.max_row - len(total_fees_per_year) - 1].height = main_sheet.row_dimensions[1].height

    # Remove excess column width and row height from the table content
    for col in range(1, overview_sheet.max_column + 1):
        col_letter = get_column_letter(col)
        overview_sheet.column_dimensions[col_letter].width = 15

    # Save the workbook with the new overview sheet
    workbook.save(output_file_path)
    print(f"Overview sheet added to {output_file_path}")


def format_dates_and_currency(output_file_path):
    # Load the existing workbook
    workbook = load_workbook(output_file_path)
    main_sheet = workbook[workbook.sheetnames[0]]  # Assuming the main sheet is the first one

    # Define date, currency, and alignment styles
    date_style = NamedStyle(name='short_date', number_format='MM/DD/YYYY')
    currency_style = NamedStyle(name='currency', number_format='$#,##0.00')
    alignment_style = Alignment(horizontal='center', vertical='top')
    header_alignment_style = Alignment(horizontal='center', vertical='center')
    bold_font = Font(bold=True)

    # Apply date style to all date columns (excluding header)
    date_columns = ['C', 'D', 'E', 'F']  # Assuming C to F are the date columns (Priority Date, Filing Date, Issued Date, Expiration Date)
    for col in date_columns:
        for cell in main_sheet[col][1:]:  # Skip the header row
            cell.style = date_style

    # Apply currency style to columns from K onwards (excluding header)
    for col in range(11, main_sheet.max_column + 1):  # Column K is the 11th column
        col_letter = get_column_letter(col)
        for cell in main_sheet[col_letter][1:]:  # Skip the header row
            cell.style = currency_style

    # Set column width to 15 for date columns and 15 for others in the main sheet
    for col in range(1, main_sheet.max_column + 1):
        col_letter = get_column_letter(col)
        if col_letter in date_columns:
            main_sheet.column_dimensions[col_letter].width = 15
        else:
            main_sheet.column_dimensions[col_letter].width = 15

    # Set row height for header row
    main_sheet.row_dimensions[1].height = 30

    # Set zoom level to 80% for the main sheet
    main_sheet.sheet_view.zoomScale = 80

    # Apply alignment and bold font to all cells (including header)
    for row in main_sheet.iter_rows():
        for cell in row:
            if cell.row == 1:  # Header row
                cell.alignment = header_alignment_style
                cell.font = bold_font
            else:
                cell.alignment = alignment_style

    # Rename the main sheet
    main_sheet.title = 'Fees per Year'

    # Format the Overview sheet
    if 'Overview' in workbook.sheetnames:
        overview_sheet = workbook['Overview']
        for cell in overview_sheet['B'][1:]:  # Assuming the numbers are in column B, skip the header
            cell.style = currency_style
        
        # Set column width to 20 for the overview sheet
        for col in range(1, overview_sheet.max_column + 1):
            col_letter = get_column_letter(col)
            overview_sheet.column_dimensions[col_letter].width = 20

        # Set row height for header row in overview sheet
        overview_sheet.row_dimensions[1].height = 30

        # Apply alignment and bold font to all cells in the overview sheet
        for row in overview_sheet.iter_rows():
            for cell in row:
                if cell.row == 1:  # Header row
                    cell.alignment = header_alignment_style
                    cell.font = bold_font
                else:
                    cell.alignment = alignment_style

    # Save the workbook with the formatting changes
    workbook.save(output_file_path)